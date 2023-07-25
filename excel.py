import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side

wb = load_workbook('excel.xlsx')
ws = wb.active
#預計要刪除ws.delete_cols(1,3,6,7,8,9,10,11,12)
ws.delete_cols(1)
ws.delete_cols(2)
#每刪掉一列都會讓index -1, 所以第二次執行delete_cols的時候都要考慮上一次刪除幾個
#(4,7)代表第四列以後連續刪除七列
ws.delete_cols(4,7)

wb.save('excel.xlsx')
border_style = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
for row in ws.iter_rows():
    for cell in row:
        cell.border = border_style

wb.save('excel.xlsx')